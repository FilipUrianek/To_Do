
# -*- coding: utf-8 -*-


import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
import openpyxl
import os
import sys

FILE_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    #your path to excel
)
HEADERS = ["Date", "Kniha", "Úklid", "Investice", "Šetření", "Práce", "Programování", "Cvičení", "SkinCare"]
TASKS = HEADERS[1:]


def ensure_workbook(path):
    if not os.path.exists(path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(HEADERS)
        wb.save(path)


def load_workbook(path):
    return openpyxl.load_workbook(path)


def find_row_for_date(ws, date_str):
    for row in ws.iter_rows(min_row=2, values_only=False):
        cell = row[0]
        if cell.value == date_str:
            return cell.row
    return None


def create_row_for_date(ws, date_str):
  
    values = [date_str] + [False] * len(TASKS)
    ws.append(values)
    return ws.max_row


def save_task_states(path, date_str, states_dict):
   
    wb = load_workbook(path)
    ws = wb.active
    row = find_row_for_date(ws, date_str)
    if row is None:
        row = create_row_for_date(ws, date_str)
   
    for i, task in enumerate(TASKS, start=2):
        val = states_dict.get(HEADERS[i-1], False)
        ws.cell(row=row, column=i, value=bool(val))
    wb.save(path)
    wb.close()


# === GUI ===
class DailyTodoApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('Denní ToDo')
        self.resizable(False, False)

    
        ensure_workbook(FILE_PATH)

      
        self.vars = {}
        for tname in TASKS:
            self.vars[tname] = tk.BooleanVar(value=False)

       
        today = datetime.now().date().isoformat()
        wb = load_workbook(FILE_PATH)
        ws = wb.active
        row = find_row_for_date(ws, today)
        if row is not None:
            for idx, tname in enumerate(TASKS, start=2):
                cell_val = ws.cell(row=row, column=idx).value
               
                self.vars[tname].set(bool(cell_val))
        wb.close()

       
        frm = ttk.Frame(self, padding=12)
        frm.grid(row=0, column=0)

        ttk.Label(frm, text=f'Dnešní datum: {today}').grid(row=0, column=0, columnspan=2, pady=(0,8))

        for i, tname in enumerate(TASKS, start=1):
            cb = ttk.Checkbutton(frm, text=tname, variable=self.vars[tname], command=self.on_change)
            cb.grid(row=i, column=0, sticky='w', pady=2)

        btn_frame = ttk.Frame(frm)
        btn_frame.grid(row=len(TASKS)+1, column=0, pady=(10,0))

        save_btn = ttk.Button(btn_frame, text='Uložit', command=self.manual_save)
        save_btn.grid(row=0, column=0, padx=6)

        close_btn = ttk.Button(btn_frame, text='Zavřít', command=self.on_close)
        close_btn.grid(row=0, column=1, padx=6)

        
        self.on_change()

        self.protocol('WM_DELETE_WINDOW', self.on_close)

    def get_states(self):
        return {name: var.get() for name, var in self.vars.items()}

    def on_change(self):

        today = datetime.now().date().isoformat()
        try:
            save_task_states(FILE_PATH, today, self.get_states())
        except Exception as e:
            messagebox.showerror('Chyba', f'Při ukládání do Excelu došlo k chybě:\n{e}')

    def manual_save(self):
        self.on_change()
        messagebox.showinfo('Uloženo', 'Stav byl uložen do Excelu.')

    def on_close(self):
   
        self.on_change()
        self.destroy()


def main():
   
    app = DailyTodoApp()
    app.mainloop()


if __name__ == '__main__':
    main()


